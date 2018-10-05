# coding=gbk
# excel����,��ָ���ļ�����"item[1]�ڶ���"��������,��������ָ���ļ���.
# ����:
# ��Ʒ����    ��Ʒ�۸�    ��Ʒ����

# - ����һ�������� readwb(wbname, sheetname=None)
# - ����û�ָ��sheetname�ʹ��û�ָ���Ĺ����� ���û��ָ���� ��active sheet;
# -  ������Ʒ�ļ۸��������(��С����)�� ���浽�ļ���;��Ʒ����:��Ʒ�۸�:��Ʒ����

import openpyxl


def readwb(wbname, sheetname=None):
    workbo = openpyxl.load_workbook(wbname)
    if not sheetname:
        sheet = workbo.active
    else:
        sheet = sheetname
    wb_info = []
    for row in sheet.rows:
        row_info = [val.value for val in row]
        wb_info.append(row_info)
    return sorted(wb_info, key=lambda item:item[2])


def save_to_excel(data, wbname, wbsheet='Sheet1'):
    workbo = openpyxl.Workbook()
    sheet = workbo.active
    for row_index,row in enumerate(data):
        for column_index,cell_value in enumerate(row):
            sheet.cell(row=row_index+1, column=column_index+1, value=cell_value)
    workbo.save(wbname)


excel_info = readwb('example.xlsx')
save_to_excel(excel_info, 'modify.xlsx')


