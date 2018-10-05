# coding=gbk
import openpyxl
# 1.����excel�ĵ�
workbo = openpyxl.load_workbook("example.xlsx") # <class 'openpyxl.workbook.workbook.Workbook'>
all_sheet = workbo.sheetnames   # ��ʾ���еĹ�����['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']
active_sheet = workbo.active    # ��ʾ���<Worksheet "Sheet4">
# 2.ѡ��Ҫ�����Ĺ�����
print(all_sheet)
print(active_sheet)
sheet = active_sheet  # workbo['Sheet1']        # ѡ��Ҫ�����Ĺ�����
title_sheet = sheet.title       # ��ȡ���������ƣ�Sheet1
print(title_sheet)
#3.�鿴ָ���ĵ�Ԫ����Ϣ
cell_b1 = sheet.cell(row=1,column=2)    # ����ָ����Ԫ����Ϣ,����һ������<Cell 'Sheet1'.B1>
value_b1 = cell_b1.value                # �鿴ָ����Ԫ�����ݣ�17

#4.��ȡ�����������������
max_row = sheet.max_row         # ��ȡ����������
max_column = sheet.max_column   # ��ȡ����������
print(max_row,max_column)
#5.�޸Ĺ�������Ϣ�͵�Ԫ����Ϣ
sheet.title = 'family_info'     # �޸Ĺ�����ı���
sheet['B1'] = 100               # �޸ĵ�Ԫ��ֵ
value=sheet.cell(1,3).value     # ��ȡ����һ�е����е�ֵ
sheet.cell(5,3).value='abc'     # �������е����е�ֵ��Ϊabc
print(value,sheet)
#6.�������е�Ԫ����Ϣ
info_row = sheet.rows               # ���ص���һ����������<generator object Worksheet._cells_by_row at 0x000001DB770DFF68>
for row in info_row:          # �����������ó�ÿһ��ֵ
    for column in row:
        print(column.value,end=' ')
    print()

#7.�����޸���Ϣ
workbo.save('example.xlsx')
# http://www.goobiao.com/
# https://segmentfault.com/a/1190000016367370